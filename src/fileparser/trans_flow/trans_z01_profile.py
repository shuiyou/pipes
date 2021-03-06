import inspect
from fileparser.trans_flow.trans_config import MAX_TITLE_NUMBER
from openpyxl.utils import get_column_letter as colref
from util.pyheaderfile import Xlsx, Xls, Csv
from fileparser.trans_flow.trans_z04_time_standardization import dttime_apply
from fileparser.trans_flow.trans_csv import TransCsv
from fileparser.trans_flow.trans_xls import TransXls
import datetime
import openpyxl
import pandas as pd
import re
import os

from logger.logger_util import LoggerUtil

logger = LoggerUtil().logger(__name__)


class TransProfile:
    """
    流水文件读取,并和属性校验
    标题行校验:检验流水文件中前30行(从第一行有数据的行开始算起)是否包含符合规范的标题行
    属性校验:包括银行名称,银行账号,户名校验
    author:汪腾飞
    created_time:20200630
    v1.0
    updated_time_v1:20200706,表头搜索到银行名将不再更改客户所填入参,而直接将该银行名存入account表中
    updated_time_v2:20200817,读取其他类型流水文件,包括xls,csv, 无法体现户名,账户信息话术统一
    updated_time_v3:20201126,流水起始日期和截止日期都进行标准化转换,即若日期格式是形如43832的数字,也将其转化为日期
                             同时户名新增模糊校验,即若姓名做了脱敏处理,只要未脱敏的部分能匹配则算匹配上
    v2.0
    updated_time_v4:20201223,
        1.不再进行户名，账号，银行名称校验，且不再更改客户所填银行名称与账号
        2.csv和xls现在均可以正常读取，不再需要另存为xlsx后再读取，且无论流水数据放在第几页都可进行识别
        3.流水标题行需要同时识别到：交易时间，交易金额，交易余额，交易对手这几列的关键字，即新增了交易对手这个列的识别
    """

    def __init__(self, file, param):
        super().__init__()
        self.file = file
        self.param = param
        self.ws = None
        self.maxcol = None
        self.maxrow = None
        self.mincol = None
        self.minrow = None
        self.title = None
        self.title_params = {}
        self.trans_data = None
        self.basic_status = True
        self.resp = {
            "resCode": "0",
            "resMsg": "成功",
            "data": {
                "warningMsg": []
            }
        }

    def process(self):
        # try:
        #     self.ws = self._load_worksheet()
        #     self.maxcol = self.ws.max_column
        #     self.maxrow = self.ws.max_row
        #     self.mincol = self.ws.min_column
        #     self.minrow = self.ws.min_row
        #     self.title = self._find_title()
        #     self.trans_title_check()
        #     if self.basic_status:
        #         self.trans_data = self._trans_data()
        # except Exception as e:
        #     logger.error(e)
        #     self.basic_status = False
        #     self.resp['resCode'] = '1'
        #     self.resp['resMsg'] = '失败'
        #     self.resp['data']['warningMsg'] = ['上传文件类型错误,仅支持csv,xls,xlsx格式文件']
        #     return
        file_type = str(self.file.filename)[-5:]
        if 'csv' in file_type:
            trans_class = TransCsv(self.file)
        elif 'xls' in file_type:
            trans_class = TransXls(self.file)
        else:
            self.basic_status = False
            self.resp['resCode'] = '1'
            self.resp['resMsg'] = '失败'
            self.resp['data']['warningMsg'] = ['上传文件类型错误,仅支持csv,xls,xlsx格式文件']
            return
        trans_class.process()
        self.basic_status = trans_class.basic_status
        self.trans_data = trans_class.trans_data
        self.resp = trans_class.resp

    # 将流水所在整个工作表读到内存中
    def _load_worksheet(self):
        logger.info("%s-----------------------%s" % (1, '_load_worksheet begin'))
        new_file = False
        if "xlsx" not in str(self.file.filename)[-5:]:
            xlsx = Xlsx()
            now_timestamp = datetime.datetime.timestamp(datetime.datetime.now())
            file_name = '%d.xlsx' % (now_timestamp * 1000)

            temp = self._guess_type(self.file, file_name)
            header_list = temp.header
            length = len(header_list)
            # 这一步是因为pyheaderfile读取文件时如果第一行存在太多空值,就会忽略掉第一个空值往后的所有列,因此需要给第一行赋值
            temp.header = [header_list[i] if header_list[i] != '' and header_list[i] is not None
                           else 'Title%d' % i for i in range(length)]
            temp.name = file_name
            # 将csv, xls文件都另存为xlsx文件
            logger.info("%s-----------------------%s" % (2, '_load_worksheet xlsx'))
            xlsx(temp)
            logger.info("%s-----------------------%s" % (3, '_load_worksheet xlsx end'))
            new_file = True
        else:
            file_name = self.file
        wb = openpyxl.load_workbook(file_name)
        ws = wb.worksheets[0]
        wb.close()
        # 将刚刚另存为的文件删除
        if new_file and os.path.exists(file_name):
            os.remove(file_name)
        logger.info("%s-----------------------%s" % (4, '_load_worksheet end'))
        return ws

    # 搜索最多前30行(从第一行有数据的地方开始搜索),查找标题行所在行,若不存在标题行则返回-1
    def _find_title(self):
        ws = self.ws
        max_cell = colref(self.maxcol) + str(self.maxrow)
        min_cell = colref(self.mincol) + str(self.minrow)
        rng = ws[min_cell:max_cell]
        min_search = min(self.maxrow - self.minrow + 1, MAX_TITLE_NUMBER)
        col_len = 0
        title = -1
        for i in range(min_search):
            temp = [str(x.value) for x in rng[i] if x.value]
            temp_len = len(temp)
            string = ''.join(temp)
            # todo 此处关键字可能会调整(不,并不会)
            if ('交易' in string and ('金额' in string or '余额' in string)) \
                    or ('日期' in string and '摘要' in string) or ('Balance' in string):
                if temp_len == self.maxcol - self.mincol + 1:
                    title = i + self.minrow
                    break
                elif temp_len > col_len:
                    col_len = temp_len
                    title = i + self.minrow
        return title

    def trans_title_check(self):
        """
        1.校验是否存在标题行,若不存在则基本校验状态设置为False,回执信息填充为解析失败
        2.检索标题行以上是否存在银行信息,若存在则将入参中的银行替换为标题行以上的银行信息
        3.检索标题行以上是否存在账号信息,若存在则对比入参和搜索到的账号信息,不匹配则基本校验状态设置为False
            回执信息填充为参数错误,若不存在则提示信息加上"无法体现账号信息"
        4.检索标题行以上是否存在户名信息,若存在则对比入参和检索到的户名信息,不匹配则基本校验状态设置为False
            回执信息填充为参数错误,若不存在则提示信息加上"无法体现户名信息"
        5.检索标题行以上是否存在查询起始和截止时间,若存在则保存在title_param中
        :return:
        """
        if self.title == -1:
            self.basic_status = False
            self.resp['resCode'] = '20'
            self.resp['resMsg'] = '解析失败'
            self.resp['data']['warningMsg'] = ['上传失败,无法找到标题行,流水文件内容有误']
            return
        elif self.title == self.minrow:
            self.resp['resCode'] = '0'
            self.resp['resMsg'] = '成功'
            self.resp['data']['warningMsg'] = ['该流水无法体现账户信息,请线下核实',
                                               '该流水无法体现户名信息,请线下核实']
            return
        else:
            check_df = self._convert_to_dataframe(self.ws, self.minrow, self.mincol, self.title - 1,
                                                  self.maxcol, False, False)
            for i in range(self.title - self.minrow):
                for j in range(self.maxcol - self.mincol + 1):
                    temp_value = check_df.iloc[i, j]
                    if pd.isnull(temp_value):
                        continue
                    self._single_cell_match(temp_value, 'raw')
                    temp_value = re.sub(r'\s', '', str(temp_value))
                    if len(temp_value) >= 5:
                        continue
                    if '账号' in temp_value:
                        if i < self.title - self.minrow - 1:
                            below_value = check_df.iloc[i+1, j]
                            if pd.notnull(below_value):
                                self._single_cell_match(below_value, 'account_no')
                        if j < self.maxcol - self.mincol:
                            right_value = check_df.iloc[i, j+1]
                            if pd.notnull(right_value):
                                self._single_cell_match(right_value, 'account_no')
                    elif '户名' in temp_value:
                        if i < self.title - self.minrow - 1:
                            below_value = check_df.iloc[i+1, j]
                            if pd.notnull(below_value):
                                self._single_cell_match(below_value, 'opponent_name')
                        if j < self.maxcol - self.mincol:
                            right_value = check_df.iloc[i, j+1]
                            if pd.notnull(right_value):
                                self._single_cell_match(right_value, 'opponent_name')
                    elif re.search(r'.*[起|开]始.*', temp_value):
                        if i < self.title - self.minrow - 1:
                            below_value = check_df.iloc[i+1, j]
                            if pd.notnull(below_value):
                                self._single_cell_match(below_value, 'start_date')
                        if j < self.maxcol - self.mincol:
                            right_value = check_df.iloc[i, j+1]
                            if pd.notnull(right_value):
                                self._single_cell_match(right_value, 'start_date')
                    elif re.search(r'.*(截止|结束|终止).*', temp_value):
                        if i < self.title - self.minrow - 1:
                            below_value = check_df.iloc[i+1, j]
                            if pd.notnull(below_value):
                                self._single_cell_match(below_value, 'end_date')
                        if j < self.maxcol - self.mincol:
                            right_value = check_df.iloc[i, j+1]
                            if pd.notnull(right_value):
                                self._single_cell_match(right_value, 'end_date')
            # 若表头所显示的银行信息与客户所填银行信息不匹配则将客户所填银行信息更新为表头所显示的银行信息
            # if self.title_params.__contains__('bank'):
            #     self.param['bankName'] = self.title_params['bank']
            if self.title_params.__contains__('account_no'):
                self.param['bankAccount'] = self.title_params['account_no'][0]
            #     if self.param['bankAccount'] not in self.title_params['account_no']:
            #         self.basic_status = False
            #         self.resp['resCode'] = '10'
            #         self.resp['resMsg'] = '参数错误'
            #         self.resp['data']['warningMsg'] = ['上传失败,流水账号信息与客户所填账号信息不匹配,请重新填写']
            #         return
            # else:
            #     self.resp['data']['warningMsg'].append('该流水无法体现账号信息,请线下核实')
            if self.title_params.__contains__('opponent_name'):
                temp_name = self.param['cusName']
                name_status = False
                for v in self.title_params['opponent_name']:
                    temp_v = re.sub(r'[*?\s]+', '.*', v)
                    temp_comp = re.compile(temp_v)
                    if re.search(temp_comp, temp_name):
                        name_status = True
                        break
                if not name_status:
                    self.basic_status = False
                    self.resp['resCode'] = '10'
                    self.resp['resMsg'] = '参数错误'
                    self.resp['data']['warningMsg'] = ['上传失败,流水户名信息与客户所填姓名不匹配,请重新填写']
                    return
            else:
                self.resp['data']['warningMsg'].append('该流水无法体现户名信息,请线下核实')
            return

    def _trans_data(self):
        df = self._convert_to_dataframe(self.ws, self.title, self.mincol, self.maxrow, self.maxcol)
        return df

    def _single_cell_match(self, cell_value, cell_type):
        """
        单个单元格匹配,可能的格式:
            1.单个关键字,如:"账号","账   号","账   号:"等
            2.单个关键字+值,如:"账号:123456","账号123456"
            3.多个关键字+值,如:"户名:张三  账号:123456"
            4.空值
        :param cell_value: 单个单元格里面的字符串
        :param cell_type: 枚举值["raw", "account_no", "opponent_name", "start_date", "end_date"]
                "raw": 所有关键字均需要查找并匹配相应关键字格式
                "account_no": 仅匹配账户格式
                "opponent_name": 仅匹配户名格式
                "start_date": 仅匹配日期格式
                "end_date": 仅匹配日期格式
        :return:
        """
        cell_value = str(cell_value).strip()
        cell_list = cell_value.split()
        if len(cell_list) == 0:
            return
        if cell_type == "raw":
            for val in cell_list:
                sub_temp = re.sub(r'[^*?:：0-9\u4e00-\u9fa5]', '', val)
                # todo 除银行外还有其他信息 如xx银行交易明细之类
                # if '银行' in sub_temp and not self.title_params.__contains__('bank'):
                #     self.title_params['bank'] = sub_temp
                # todo 关键字卡号/卡号被空格隔开
                if '账号' in sub_temp or '卡号' in sub_temp:
                    acc_temp = re.sub(r'[*?]', '', sub_temp)
                    acc_temp = re.search(r'[3-9]\d{12,18}', acc_temp)
                    if acc_temp is not None:
                        if self.title_params.__contains__('account_no'):
                            self.title_params['account_no'].append(acc_temp.group(0))
                        else:
                            self.title_params['account_no'] = [acc_temp.group(0)]
                # todo 关键字户名/户名和姓名被空格隔开
                if '户名' in sub_temp or '姓名' in sub_temp:
                    name_temp = re.search(r'(?<=[\u4e00-\u9fa5][:：])[*?\u4e00-\u9fa5]{2,4}', sub_temp)
                    if name_temp is not None:
                        if self.title_params.__contains__('opponent_name'):
                            self.title_params['opponent_name'].append(name_temp.group(0))
                        else:
                            self.title_params['opponent_name'] = [name_temp.group(0)]
                # todo 关键字 本次查询时间段
                sub_temp = re.sub(r'[*?]', '', sub_temp)
                if re.search(r'[起|开]始', sub_temp):
                    start_temp = re.search(r'^20([01]\d|20)(0[1-9]|1[012])(0[1-9]|[12]\d|3[01])|^4\d{4}', sub_temp)
                    if start_temp is not None:
                        self.title_params['start_date'] = format(dttime_apply(start_temp.group(0)), '%Y-%m-%d')
                if re.search(r'(截止|结束|终止)', sub_temp):
                    end_temp = re.search(r'^20([01]\d|20)(0[1-9]|1[012])(0[1-9]|[12]\d|3[01])|^4\d{4}', sub_temp)
                    if end_temp is not None:
                        self.title_params['end_date'] = format(dttime_apply(end_temp.group(0)), '%Y-%m-%d')
        else:
            cell_value = ''.join(cell_list)
            cell_value = re.sub(r'[^*?\d\u4e00-\u9fa5]', '', cell_value)
            if cell_type == 'account_no':
                if re.match(r'[3-9]\d{12,18}', cell_value):
                    if self.title_params.__contains__('account_no'):
                        self.title_params['account_no'].append(cell_value)
                    else:
                        self.title_params['account_no'] = [cell_value]
            elif cell_type == 'opponent_name':
                if re.match(r'[*?\u4e00-\u9fa5]{2,}', cell_value):
                    if self.title_params.__contains__('opponent_name'):
                        self.title_params['opponent_name'].append(cell_value)
                    else:
                        self.title_params['opponent_name'] = [cell_value]
            elif cell_type == 'start_date':
                cell_value = re.sub(r'[*?]', '', cell_value)
                if re.match(r'^20([01]\d|20)(0[1-9]|1[012])(0[1-9]|[12]\d|3[01])|^4\d{4}', cell_value):
                    self.title_params['start_date'] = format(dttime_apply(cell_value), '%Y-%m-%d')
            elif cell_type == 'end_date':
                cell_value = re.sub(r'[*?]', '', cell_value)
                if re.match(r'^20([01]\d|20)(0[1-9]|1[012])(0[1-9]|[12]\d|3[01])|^4\d{4}', cell_value):
                    self.title_params['end_date'] = format(dttime_apply(cell_value), '%Y-%m-%d')
        return

    @staticmethod
    def _convert_to_dataframe(worksheet, min_row, min_col, max_row, max_col, index=False, column=True):
        """
        将工作表的对应区域转化为二维数据表
        :param worksheet: 工作表数据
        :param min_row: 需要转化的数据区域最小行数
        :param min_col: 需要转化的数据区域最小列数
        :param max_row: 需要转化的数据区域最大行数
        :param max_col: 需要转化的数据区域最大列数
        :param index: 是否有行标签列,默认不包含行标签列
        :param column: 是否有列标签行,默认包含列标签行
        :return: pd.DataFrame
        """
        min_ref = colref(min_col) + str(min_row)
        max_ref = colref(max_col) + str(max_row)
        rng = worksheet[min_ref:max_ref]
        if index:
            index_start = 1
            rows = [r[0].value for r in rng]
            if column:
                rows.pop(0)
        else:
            index_start = 0
            rows = None
        if column:
            column_start = 1
            cols = [c.value for c in rng[0]]
            col_len = len(cols)
            for c_index in range(col_len):
                if pd.isna(cols[c_index]):
                    cols[c_index] = 'Na%d' % c_index
                else:
                    cols[c_index] = re.sub(r'[\s| ]', '', str(cols[c_index])).strip()
            if index:
                cols.pop(0)
        else:
            column_start = 0
            cols = None
        data = [[x.value for x in y[index_start:]] for y in rng[column_start:]]
        df = pd.DataFrame(data=data, index=rows, columns=cols)
        return df

    def _guess_type(self, file, filename, **kwargs):
        """ Utility function to call classes based on filename extension.
        Just usefull if you are reading the file and don't know file extension.
        You can pass kwargs and these args are passed to class only if they are
        used in class.
        """

        extension = os.path.splitext(file.filename.replace('"', ""))[1]
        case = {'.xls': Xls,
                '.xlsx': Xlsx,
                '.csv': Csv}
        if extension and case.get(extension.lower()):
            low_extension = extension.lower()
            new_kwargs = dict()
            class_name = case.get(low_extension)
            class_kwargs = inspect.getargspec(class_name.__init__).args[1:]
            for kwarg in kwargs:
                if kwarg in class_kwargs:
                    new_kwargs[kwarg] = kwargs[kwarg]
            file.save(filename)
            return case.get(low_extension)(filename, **new_kwargs)
        else:
            raise Exception('No extension found')